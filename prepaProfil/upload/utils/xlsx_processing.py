from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from .fill import fill_emplacements


def save_excel(rows, stock_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Ref", "Qtes", "Commande", "Adresses (qtes)"])
    for row in rows:
        ws.append([row["date"], row["code"], row["qty"], row["cmd"], ""])
    fill_emplacements(ws, stock_file=stock_file)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                  min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.border = thin_border
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
