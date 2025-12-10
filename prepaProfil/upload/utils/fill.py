import msoffcrypto
import io
import os
from openpyxl import load_workbook


def fill_emplacements(ws_day, stock_file, password="alphacan"):

    decrypted = io.BytesIO()
    with open(stock_file, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)

    wb_stock = load_workbook(decrypted, data_only=True)
    if "SDM" not in wb_stock.sheetnames:
        print("ERREUR : la feuille 'SDM' n'existe pas. Feuilles :",
              wb_stock.sheetnames)
        return
    ws_sdm = wb_stock["SDM"]

    header_row = None
    headers = []
    for i, row in enumerate(
            ws_sdm.iter_rows(min_row=1, max_row=10, values_only=True),
            start=1):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header_row = i
            headers = [str(c).strip().lower() if c is not None else "" for c in
                       row]
            break
    if header_row is None:
        print("ERREUR : impossible de détecter la ligne d'en-tête dans SDM.")
        return

    def find_idx(names):
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    idx_adresse = find_idx(["adresse", "address"]) or 0
    idx_ref = find_idx(["ref", "réf", "reference", "référence"]) or 1
    idx_couleur = find_idx(["couleur", "color"]) or 2
    idx_qte = find_idx(["qte", "qty", "quantite", "quantité"]) or 6

    stock_dict = {}
    for row in ws_sdm.iter_rows(min_row=header_row + 1, values_only=True):
        adresse = str(row[idx_adresse]).strip() if row[idx_adresse] else ""
        ref_sdm = str(row[idx_ref]).strip().upper() if row[idx_ref] else None
        couleur_sdm = str(row[idx_couleur]).strip().upper() if row[
            idx_couleur] else None
        qte = row[idx_qte] if len(row) > idx_qte else 0

        if not ref_sdm or not couleur_sdm:
            continue

        key = (ref_sdm, couleur_sdm)
        stock_dict.setdefault(key, []).append({"adresse": adresse, "qte": qte})

    for row in ws_day.iter_rows(min_row=2):
        full_ref_cell = row[1].value
        if not full_ref_cell or len(str(full_ref_cell)) < 4:
            row[4].value = ""
            continue

        full_ref = str(full_ref_cell).strip().upper()
        code_ref = full_ref[:-3].strip().upper()
        couleur = full_ref[-3:].strip().upper()

        found_entries = stock_dict.get((code_ref, couleur), [])

        if not found_entries:
            for (ref_sdm, couleur_sdm), entries in stock_dict.items():
                if couleur_sdm == couleur and (
                        ref_sdm == code_ref or ref_sdm.startswith(
                        code_ref) or code_ref in ref_sdm or ref_sdm.endswith(
                        code_ref)):
                    found_entries.extend(entries)

        if found_entries:
            seen = []
            for e in found_entries:
                addr = (e.get("adresse") or "").strip()
                qty = e.get("qte") or 0
                if addr and (addr, qty) not in seen:
                    seen.append((addr, qty))
            row[4].value = " | ".join(f"{a} ({q})" for a, q in seen)
        else:
            row[4].value = ""
